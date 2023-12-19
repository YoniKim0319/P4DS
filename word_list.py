from openpyxl import load_workbook
from konlpy.tag import Okt
from matplotlib import font_manager as fm

# 엑셀 파일에서 데이터 읽어오기
wordcloud_data = 'storyline_update2.xlsx'
read_xlsx = load_workbook(wordcloud_data, data_only=True)
read_sheet = read_xlsx.active
name_col = read_sheet['A1:A1128']

# 줄거리를 행 별로 리스트로 만들기
storyline = [cell.value for row in read_sheet.iter_rows(min_row=2, max_row=1128, min_col=1, max_col=1) for cell in row]

# 한글 폰트 설정
font_path = 'C:\\Users\\USER\\AppData\\Local\\Microsoft\\Windows\\Fonts\\NanumGothicBold.ttf'
font_prop = fm.FontProperties(fname=font_path).get_name()

# Matplotlib에서 한글 폰트 설정
import matplotlib
matplotlib.rc('font', family=font_prop)
matplotlib.rcParams['axes.unicode_minus'] = False

# Okt tokenizer 불러오기
okt = Okt()

# 리스트를 저장할 변수 초기화
text_for_wordcloud_list = []

# 각 행의 줄거리를 사용하여 WordCloud 생성 및 시각화
for row_num, row in enumerate(read_sheet.iter_rows(min_row=2, max_row=1128, min_col=1, max_col=1), start=2):
    storyline = row[0].value

    # 텍스트 데이터를 전처리
    text_data = ' '.join(storyline)

    # Okt를 사용하여 명사와 형용사를 추출하기
    tokens = okt.pos(text_data, stem=True)
    filtered_tokens = [word for word, pos in tokens if pos in ['Noun', 'Adjective']]

    # WordCloud에 사용할 텍스트로 필터링된 토큰을 결합
    text_for_wordcloud = ' '.join(filtered_tokens)

    # 리스트에 추가
    text_for_wordcloud_list.append(text_for_wordcloud)

print(text_for_wordcloud_list)

import openpyxl

# 엑셀 파일 생성 또는 열기
excel_file_path = 'drama_complete.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)

# 원하는 시트 선택 (예: Sheet1)
sheet = workbook['Sheet1']

# 파이썬 리스트 데이터 : 

# 데이터를 G열의 2행부터 순차적으로 쓰기
column_G = sheet['G']
for i, value in enumerate(text_for_wordcloud_list, start=2):
    column_G[i - 1].value = value

# 엑셀 파일 저장
workbook.save(excel_file_path)
