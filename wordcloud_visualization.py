import openpyxl
from konlpy.tag import Komoran
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from matplotlib import font_manager as fm

# 엑셀 파일 경로 지정
excel_file_path = 'storyline_update2.xlsx'  # 실제 파일 경로로 변경해야 합니다.

# 엑셀 파일 열기
workbook = openpyxl.load_workbook(excel_file_path)

# 원하는 시트 선택 (예: 첫 번째 시트 선택)
sheet = workbook.active

# 한글 폰트 설정
font_path = 'C:\\Users\\USER\\AppData\\Local\\Microsoft\\Windows\\Fonts\\NanumGothicBold.ttf'
font_prop = fm.FontProperties(fname=font_path).get_name()
import matplotlib
matplotlib.rc('font', family=font_prop)

# 유니코드 깨짐현상 해결
matplotlib.rcParams['axes.unicode_minus'] = False

# KoNLPy의 Komoran 객체 생성
komoran = Komoran()

# 명사와 형용사 추출
nouns_and_adjectives_per_row_dict = {}
for row in range(1, sheet.max_row + 1):
    text = sheet.cell(row=row, column=1).value
    if text:
        pos_tags = komoran.pos(text)
        # 명사와 형용사만 추출하여 딕셔너리에 저장
        extracted_words = [word for word, pos in pos_tags if pos in ['NN', 'NNG', 'NNP', 'VA', 'VAX', 'XR']]
        nouns_and_adjectives_per_row_dict[row] = extracted_words

# 딕셔너리 값들의 단어들을 모두 리스트로 합치기
all_words = [word for words in nouns_and_adjectives_per_row_dict.values() for word in words]

# 리스트의 단어들을 공백으로 구분된 문자열로 변환
text_for_wordcloud = ' '.join(all_words)

# WordCloud 생성 - Specify Korean font directly
wordcloud = WordCloud(width=800, height=400, background_color='white', font_path=font_path).generate(text_for_wordcloud)

# Matplotlib을 사용하여 WordCloud 출력
plt.figure(figsize=(10, 5))
plt.imshow(wordcloud, interpolation='bilinear')
plt.axis('off')
plt.show()
