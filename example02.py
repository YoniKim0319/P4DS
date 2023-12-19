from openpyxl import load_workbook
from konlpy.tag import Okt
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from matplotlib import font_manager as fm

# 엑셀 파일에서 데이터 읽어오기
wordcloud_data = 'storyline_update2.xlsx'
read_xlsx = load_workbook(wordcloud_data, data_only=True)
read_sheet = read_xlsx.active
name_col = read_sheet['A1:A1128']

# 줄거리를 행 별로 리스트로 만들기
storyline = [cell.value for row in read_sheet.iter_rows(min_row=2, max_row=1128, min_col=1, max_col=1) for cell in row]

# 한글 폰트 설정
font_path = 'C:\\Users\\USER\\AppData\\Local\\Microsoft\\Windows\\Fonts\\NanumGothicBold.ttf'
font_prop = fm.FontProperties(fname=font_path).get_name()
import matplotlib
matplotlib.rc('font', family=font_prop)

# 유니코드 깨짐현상 해결
matplotlib.rcParams['axes.unicode_minus'] = False

# Okt tokenizer 불러오기
okt = Okt()

# 각 행의 줄거리를 사용하여 WordCloud 생성 및 시각화
for row_num, row in enumerate(read_sheet.iter_rows(min_row=2, max_row=1128, min_col=1, max_col=1), start=2):
    storyline = row[0].value

    # 텍스트 데이터를 전처리
    text_data = ' '.join(storyline)

    # Okt를 사용하여 명사와 형용사를 추출하기
    tokens = okt.pos(text_data, stem=True)
    filtered_tokens = [word for word, pos in tokens if pos in ['Noun', 'Adjective']]

    # WordCloud에 사용할 텍스트로 필터링된 토큰을 결합
    text_for_wordcloud = ' '.join(filtered_tokens)

    # wordcloud 생성
    wordcloud = WordCloud(
        font_path=font_path,
        background_color='white',
        width=800,
        height=400,
        max_words=100,  # 필요에 따라 조절
    ).generate(text_for_wordcloud)

    # 시각화
    plt.figure(figsize=(10, 5))
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis('off')
    plt.title(f'WordCloud for Row {row_num}')
    plt.show()


#-----------------------------------------------------------#
# 기존 엑셀의 H열에 2행부터 기입
column_to_update = 'H'
start_row = 2

for index, value in enumerate(text_for_wordcloud) :
    read_sheet[f'{column_to_update}{start_row + index}'] = value

#기존의 엑셀 파일에 저장
read_xlsx.save(capstone_drama_no_fail)